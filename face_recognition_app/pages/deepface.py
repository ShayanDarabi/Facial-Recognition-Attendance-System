import streamlit as st
from PIL import Image
import base64
import cv2
import numpy as np
from PIL import ImageOps
import imutils
import urllib.request
from matplotlib import pyplot
from numpy import asarray
from gfpgan import GFPGANer
from deepface import DeepFace
import time
import io
import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
from mtcnn.mtcnn import MTCNN
from streamlit_js_eval import streamlit_js_eval

# Set page configuration
st.set_page_config(
    page_title="Facial Recognition Attendance App",
    layout="centered",
)

# Custom CSS to hide the Streamlit header & footer
st.markdown("""
    <style>
        /* Hide Streamlit's default menu and footer */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# Function to extract all faces from an image and return them as in-memory objects
def extract_faces_from_image(image, required_size=(224, 224)):
    # Convert the uploaded file to a numpy array
    pixels = asarray(image)
    # Create the detector, using default weights
    detector = MTCNN()
    # Detect faces in the image
    results = detector.detect_faces(pixels)
    faces = []
    face_buffers = []  # List to store in-memory face images

    # Loop through all detected faces
    for i, result in enumerate(results):
        # Extract the bounding box for each face
        x1, y1, width, height = result['box']
        x2, y2 = x1 + width, y1 + height
        # Extract the face
        face = pixels[y1:y2, x1:x2]
        # Resize the face to the model's required size
        face_image = Image.fromarray(face)
        face_image = face_image.resize(required_size)
        face_array = asarray(face_image)
        faces.append(face_array)

        # Save the face into an in-memory buffer
        buffer = io.BytesIO()
        face_image.save(buffer, format="JPEG")
        buffer.seek(0)
        face_buffers.append(buffer)  # Add the buffer to the list

    return faces, face_buffers

# Function to enhance an image using GFPGAN
def enhance_faces_with_gfpgan(faces, gfpgan_model):
    enhanced_faces = []
    for face in faces:
        _, _, enhanced_face = gfpgan_model.enhance(face, has_aligned=False, only_center_face=False)
        enhanced_faces.append(enhanced_face)
    return enhanced_faces
    

# Initialize GFPGANer
@st.cache_resource
def initialize_gfpgan():
    model_path = "/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/GFPGAN/experiments/pretrained_models/GFPGANv1.3.pth"  
    return GFPGANer(
        model_path=model_path,
        upscale=2,
        arch='clean',
        channel_multiplier=2,
        bg_upsampler=None
    )
    
# Function to match faces using DeepFace
def match_faces_with_deepface(enhanced_faces, db_path, model_name='Facenet512', detector_backend='mtcnn'):
    student_names = []
    for i, face in enumerate(enhanced_faces):
        # Save enhanced face temporarily to a file
        temp_file_path = f"temp_face_{i}.jpg"
        Image.fromarray(face).save(temp_file_path)

        # Use DeepFace to find the match
        try:
            model = DeepFace.find(
                img_path=temp_file_path,
                db_path=db_path,
                enforce_detection=True,
                model_name=model_name,
                detector_backend=detector_backend
            )
            matched_image_path = model[0]['identity'].values[0]
            # Extract the base name without extension
            student_name = os.path.basename(matched_image_path)
            student_name_without_extension = os.path.splitext(student_name)[0]
            # Remove trailing digits using a regular expression
            student_name_cleaned = re.sub(r'\d+$', '', student_name_without_extension)
            student_names.append(student_name_cleaned)
        except Exception as e:
            student_names.append("Unknown")
        finally:
            os.remove(temp_file_path)  # Clean up temporary file
    return student_names

# Function to extract the date an image was captured
def get_date_taken(path):
    exif = Image.open(path)._getexif()
    if not exif:
        raise Exception('Image {0} does not have EXIF data.'.format(path))
    return exif[36867]

# Function to save attendance to an Excel file
def save_attendance_to_excel(student_names, excel_path="Attendance.xlsx", sheet_name=None):
    # Load or create the workbook
    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()

    # Check if sheet exists, otherwise create it
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(["Student Name"])  # Add header

    # Append student names to the sheet
    for name in student_names:
        sheet.append([name])

    # Save the workbook
    workbook.save(excel_path)


# Background image path
background_image_path = "/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/Background Images/UT Entrance.jpg"
logo_image_path = "/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/Background Images/University_of_Tehran_logo.png"

# Function to set the background image and ensure it covers the whole page
def set_background():
    with open(background_image_path, "rb") as image_file:
        encoded_image = base64.b64encode(image_file.read()).decode()
    background_style = f"""
    <style>
    .stApp {{
        background-image: url('data:image/jpeg;base64,{encoded_image}');
        background-size: cover;
        background-position: center;
    }}
    .input-overlay {{
        background: rgba(0, 0, 0, 0.5); /* Black with 50% transparency */
        color: white;
        padding: 5px; /* Reduced padding */
        border-radius: 10px;
        margin-bottom: -10px; /* Reduced bottom margin */
        text-align: center;
    }}
    .stTextInput, .stFileUploader {{
        margin-top: -15px; /* Reduce space above inputs */
    }}
    </style>
    """
    st.markdown(background_style, unsafe_allow_html=True)

# Function to encode image to base64
def image_to_base64(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()


# Set the background image
set_background()

# Add logo to the header and center it
logo_base64 = image_to_base64(logo_image_path)
st.markdown(
    f"""
    <style>
    .logo-container {{
        text-align: center;
    }}
    </style>
    <div class="logo-container">
        <img src="data:image/jpeg;base64,{logo_base64}" width="150">
    </div>
    """,
    unsafe_allow_html=True,
)

# App title
st.title("Facial Recognition Attendance App")
gfpgan_model = initialize_gfpgan()

# Upload image option with a transparent background

st.markdown(
    """
    <div class="input-overlay">
        Facial Recognition via Uploading Photos
    </div>
    """,
    unsafe_allow_html=True,
)

# ----- TOGGLE SWITCH USING STREAMLIT -----
enhance_faces_toggle = st.toggle("Enhance Faces with GFPGAN", 
                                 value=True,
                                help="""GFPGAN (Generative Facial Prior GAN) is used to enhance facial details, improving quality and restoring degraded images.
                                Check this out for more information: https://github.com/TencentARC/GFPGAN""")

# ----- USER OPTIONS FOR MODEL & DETECTOR -----
st.markdown("<div class='input-overlay'>Select Face Recognition Model</div>", unsafe_allow_html=True)
model_name = st.selectbox(
    "",
    options=["Facenet512", "Facenet", "VGG-Face"],
    index=0,  # Default to Facenet512
    help="Choose the deep learning model for face recognition. Facenet512 is more accurate but slower."
)

st.markdown("<div class='input-overlay'>Select Face Detector Backend</div>", unsafe_allow_html=True)
detector_backend = st.selectbox(
    "",
    options=["mtcnn", "retinaface", "ssd"],
    index=0,  # Default to mtcnn
    help="Select the face detection backend. MTCNN and RetinaFace are more accurate, and SSD is lightweight and faster."
)

uploaded_file = st.file_uploader("", type=["jpg", "jpeg", "png"])  # Empty label handled by overlay
if uploaded_file is not None:
    # Save uploaded file temporarily to extract EXIF data
    temp_file_path = "temp_uploaded_image.jpg"
    with open(temp_file_path, "wb") as temp_file:
        temp_file.write(uploaded_file.getvalue())

    # Get the capture date of the image
    try:
        capture_date = get_date_taken(temp_file_path)
        if capture_date is None:
            capture_date = datetime.now().strftime("%Y-%m-%d")
            st.write(f"Capture date not found. Using current date: {capture_date}")
        else:
            st.write(f"Image captured on: {capture_date}")
            # Replace invalid characters in the capture date for Excel sheet name
            capture_date = capture_date.replace(":", "-").split(" ")[0]  # Convert to 'YYYY-MM-DD'
    except Exception as e:
        capture_date = datetime.now().strftime("%Y-%m-%d")
        # Replace invalid characters in the capture date for Excel sheet name
        capture_date = capture_date.replace(":", "-").split(" ")[0]  # Convert to 'YYYY-MM-DD'
        st.write(f"Error reading capture date. Using current date: {capture_date}")
    finally:
        os.remove(temp_file_path)
        
    # Load the uploaded image
    image = Image.open(uploaded_file)
    st.image(image, caption="Uploaded Image", use_container_width=True)
    st.write("Facial recognition will process the uploaded image!")

    # Image Processing Logic
    if "uploaded_file" in locals() and uploaded_file is not None:
        with st.spinner("Extracting faces..."):
            faces, _ = extract_faces_from_image(Image.open(uploaded_file))
            
    # ----- FACE ENHANCEMENT (If toggled ON) -----
        if enhance_faces_toggle:
            st.write("#### Enhancement Mode:", "🟢 **ON**")
            with st.spinner("Enhancing faces with GFPGAN..."):
                enhanced_faces = enhance_faces_with_gfpgan(faces, gfpgan_model)
        else:
            st.write("#### Enhancement Mode:", "🔴 **OFF**")
            enhanced_faces = faces  # Skip enhancement if toggle is OFF
            
    # Match faces and get student names
    db_path = "/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/FR APP DB"
    with st.spinner("Matching faces with DeepFace..."):
        matched_faces = match_faces_with_deepface(enhanced_faces, db_path, model_name, detector_backend=detector_backend)

    # Extract student names and save attendance
    student_names = [name for name in matched_faces]
    save_attendance_to_excel(student_names, excel_path="Attendance.xlsx", sheet_name=capture_date)

# -------- Printing Results ONLY IF IMAGE WAS UPLOADED --------
# Function to resize images to a reasonable size
def resize_image(image, size=(150, 150)):  # Adjust the size as needed
    return image.resize(size, Image.LANCZOS)

import base64

# Function to convert images to Base64 format for embedding in HTML
def image_to_base64(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# Convert the left and right wing images
left_wing_base64 = image_to_base64("/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/Background Images/left wing.png")
right_wing_base64 = image_to_base64("/mnt/d/University/Master/1st Sem/Data Mining Algorithms and Applications/Face Recognition/Background Images/right wing.png")


if "student_names" in locals() and student_names:
    st.write("## Attendance Result")
    num_cols = 4  # Number of students per row
    rows = [list(zip(student_names, enhanced_faces))[i:i + num_cols] for i in range(0, len(student_names), num_cols)]  # Group students into rows

    for row in rows:
        cols = st.columns(len(row))  # Create dynamic columns

        for col, (name, face_buffer) in zip(cols, row):
            face_image = Image.fromarray(face_buffer)  # Convert NumPy array to PIL Image
            resized_face = resize_image(face_image)  # Resize face for consistent display

            with col:
                st.image(resized_face, width=150)  # Fixed size
                st.markdown(
                    f"""
                    <div style="display: flex; align-items: center; justify-content: center; width: 150px;">
                        <img src="data:image/png;base64,{left_wing_base64}" style="height: 40px; margin-right: 5px;">
                        <span style="text-align: center; font-size: 16px; font-weight: bold; 
                        color: #1C39BB; background-color: white; padding: 5px 10px; border-radius: 5px;">
                            {name}
                        </span>
                        <img src="data:image/png;base64,{right_wing_base64}" style="height: 40px; margin-left: 5px;">
                    </div>
                    """,
                    unsafe_allow_html=True
                )
elif "student_names" in locals():
    st.write("## Attendance Result")
    st.write("No matched faces found!")